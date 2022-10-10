import openpyxl

def get_row_count(file, sheet):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name(sheet)
    return worksheet.cell(row=reading)


def get_colum_count(file, sheet):

def get_colum_count(file, sheet):

def write_data(file,sheet, writing_row, writing_colum, data):
    workbook = openpyxl.load_workbook(file)